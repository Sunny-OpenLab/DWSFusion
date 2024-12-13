# -*-coding:utf-8 -*-

"""
# File       : __init__.py.py
# Time       ：2024/4/12 18:01
# Author     ：Qiang42
# version    ：python 3.8
# Description：
"""
import io
import torch
import torch.nn.functional as F
from math import exp

from PIL import Image
from Fusion_metric_python.Metric_Python.Metric import *
from natsort import natsorted
from tqdm import tqdm
import os
import pandas as pd
import tarfile
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def gaussian(window_size, sigma):
    gauss = torch.Tensor([exp(-(x - window_size//2)**2/float(2*sigma**2)) for x in range(window_size)])
    return gauss/gauss.sum()

def create_window(window_size, channel=1):
    _1D_window = gaussian(window_size, 1.5).unsqueeze(1)
    _2D_window = _1D_window.mm(_1D_window.t()).float().unsqueeze(0).unsqueeze(0)
    window = _2D_window.expand(channel, 1, window_size, window_size).contiguous()
    return window

def ssim(img1, img2, window_size=11, window=None, size_average=True, full=False, val_range=None):
    if val_range is None:
        if torch.max(img1) > 128:
            max_val = 255
        else:
            max_val = 1

        if torch.min(img1) < -0.5:
            min_val = -1
        else:
            min_val = 0
        L = max_val - min_val
    else:
        L = val_range

    padd = 0
    (_, channel, height, width) = img1.size()
    if window is None:
        real_size = min(window_size, height, width)
        window = create_window(real_size, channel=channel).to(img1.device)
    else:
        window = window.to(img1.device)
    mu1 = F.conv2d(img1, window, padding=padd, groups=channel)
    mu2 = F.conv2d(img2, window, padding=padd, groups=channel)

    mu1_sq = mu1.pow(2)
    mu2_sq = mu2.pow(2)
    mu1_mu2 = mu1 * mu2

    sigma1_sq = F.conv2d(img1 * img1, window, padding=padd, groups=channel) - mu1_sq
    sigma2_sq = F.conv2d(img2 * img2, window, padding=padd, groups=channel) - mu2_sq
    sigma12 = F.conv2d(img1 * img2, window, padding=padd, groups=channel) - mu1_mu2

    C1 = (0.01 * L) ** 2
    C2 = (0.03 * L) ** 2

    v1 = 2.0 * sigma12 + C2
    v2 = sigma1_sq + sigma2_sq + C2
    cs = torch.mean(v1 / v2)  # contrast sensitivity

    ssim_map = ((2 * mu1_mu2 + C1) * v1) / ((mu1_sq + mu2_sq + C1) * v2)

    if size_average:
        ret = ssim_map.mean()
    else:
        ret = ssim_map.mean(1).mean(1).mean(1)

    if full:
        return ret, cs
    return ret


def msssim(img1, img2, window_size=11, size_average=True, val_range=None, normalize=False):
    device = img1.device
    weights = torch.FloatTensor([0.0448, 0.2856, 0.3001, 0.2363, 0.1333]).to(device)
    levels = weights.size()[0]
    mssim = []
    mcs = []
    for _ in range(levels):
        sim, cs = ssim(img1, img2, window_size=window_size, size_average=size_average, full=True, val_range=val_range)
        mssim.append(sim)
        mcs.append(cs)

        img1 = F.avg_pool2d(img1, (2, 2))
        img2 = F.avg_pool2d(img2, (2, 2))

    mssim = torch.stack(mssim)
    mcs = torch.stack(mcs)

    if normalize:
        mssim = (mssim + 1) / 2
        mcs = (mcs + 1) / 2

    pow1 = mcs ** weights
    pow2 = mssim ** weights
    output = torch.prod(pow1[:-1] * pow2[-1])
    return output


# Classes to re-use window
class SSIM(torch.nn.Module):
    def __init__(self, window_size=11, size_average=True, val_range=None):
        super(SSIM, self).__init__()
        self.window_size = window_size
        self.size_average = size_average
        self.val_range = val_range

        self.channel = 1
        self.window = create_window(window_size)

    def forward(self, img1, img2):
        (_, channel, _, _) = img1.size()

        if channel == self.channel and self.window.dtype == img1.dtype:
            window = self.window
        else:
            window = create_window(self.window_size, channel).to(img1.device).type(img1.dtype)
            self.window = window
            self.channel = channel

        return ssim(img1, img2, window=window, window_size=self.window_size, size_average=self.size_average)

class evaluation_main():
    def __init__(self, args):
        super(evaluation_main, self).__init__()
        self.args = args
    def write_excel(self, excel_name='metric.xlsx', worksheet_name='VIF', column_index=0, data=None):
        try:
            workbook = load_workbook(excel_name)
        except FileNotFoundError:
            workbook = Workbook()
        if worksheet_name in workbook.sheetnames:
            worksheet = workbook[worksheet_name]
        else:
            worksheet = workbook.create_sheet(title=worksheet_name)

        column = get_column_letter(column_index + 1)
        for i, value in enumerate(data):
            cell = worksheet[column + str(i + 1)]
            cell.value = value
        workbook.save(excel_name)

    def evaluation_one(self, tar_path, item, f_name):
        f_img = Image.open(f_name).convert('L')
        with tarfile.open(tar_path, 'r') as tar:
            members = [m for m in tar.getmembers() if m.isfile()]
            inf_member = next((m for m in members if 'ir/' in m.name and item in m.name), None)
            vis_member = next((m for m in members if 'vi/' in m.name and item in m.name), None)
            if inf_member is None or vis_member is None:
                raise FileNotFoundError("未找到图像文件。")
            ir_image_data = tar.extractfile(inf_member).read()
            vi_image_data = tar.extractfile(vis_member).read()

        ir_img = Image.open(io.BytesIO(ir_image_data)).convert('L')
        vi_img = Image.open(io.BytesIO(vi_image_data)).convert('L')
        f_img_int = np.array(f_img).astype(np.int32)
        f_img_double = np.array(f_img).astype(np.float32)
        ir_img_int = np.array(ir_img).astype(np.int32)
        ir_img_double = np.array(ir_img).astype(np.float32)
        vi_img_int = np.array(vi_img).astype(np.int32)
        vi_img_double = np.array(vi_img).astype(np.float32)
        MI = MI_function(ir_img_int, vi_img_int, f_img_int, gray_level=256)

        SD = SD_function(f_img_double)
        VIF = VIF_function(ir_img_double, vi_img_double, f_img_double)
        SSIM = SSIM_function(ir_img_double, vi_img_double, f_img_double)
        Qabf = Qabf_function(ir_img_double, vi_img_double, f_img_double)
        PSNR = PSNR_function(ir_img_double, vi_img_double, f_img_double)

        return MI, SD, VIF, SSIM, Qabf, PSNR

    def evaluation_main(self, args):
        result_root = args.test_save_path
        test_data_root = args.test_dataset_path
        Method = args.model
        for dataset_name in os.listdir(result_root):
            if not os.path.isdir(os.path.join(result_root, dataset_name)) or "fuse_" in dataset_name:
                continue
            with_mean = True
            MI_list = []
            SD_list = []
            VIF_list = []
            SSIM_list = []
            Qabf_list = []
            PSNR_list = []
            filename_list = [dataset_name]
            result_dict = {
                'MI': MI_list,
                'SD': SD_list,
                'VIF': VIF_list,
                'SSIM': SSIM_list,
                'Qabf': Qabf_list,
                'PSNR': PSNR_list,
            }
            tar_path = os.path.join(test_data_root, dataset_name+'.tar')
            f_dir = os.path.join(result_root, dataset_name)
            save_dir = result_root
            os.makedirs(save_dir, exist_ok=True)
            metric_save_name = os.path.join(save_dir, 'metric_{}_{}.xlsx'.format(dataset_name, Method))

            with tarfile.open(tar_path, 'r') as tar:
                ir_members = [m.name for m in tar.getmembers() if m.isfile()]
                inf_members = [m for m in ir_members if 'ir/' in m]
                inf_members.sort(key=lambda m: os.path.basename(m))
                ir_files = [m[len(dataset_name) + 4:] for m in inf_members]
                filelist = natsorted(ir_files)
            eval_bar = tqdm(filelist, leave=False, delay=1)
            for _, item in enumerate(eval_bar):
                f_name = os.path.join(f_dir, item)
                MI, SD, VIF, SSIM, Qabf, PSNR = self.evaluation_one(tar_path, item, f_name)
                MI_list.append(MI)
                SD_list.append(SD)
                VIF_list.append(VIF)
                SSIM_list.append(SSIM)
                Qabf_list.append(Qabf)
                PSNR_list.append(PSNR)
                filename_list.append(item)
                eval_bar.set_description("{} | {}".format(dataset_name, item))
            if with_mean:
                for key, value in result_dict.items():
                    value.insert(0, np.mean(value))
                filename_list.insert(1, 'mean')
                for key, value in result_dict.items():
                    value.insert(1, np.std(value))
                filename_list.insert(2, 'std')
            for key, value in result_dict.items():
                value.insert(0, '{}'.format(key))
            self.write_excel(metric_save_name, 'all', 0, filename_list)
            for i, (key, value) in enumerate(result_dict.items()):
                self.write_excel(metric_save_name, 'all', i + 1, value)
            excel_file_path = metric_save_name
            csv_data = pd.read_excel(excel_file_path, sheet_name="all")
            csv_file_name = os.path.join(save_dir, 'metric_{}_{}.csv'.format(dataset_name, Method))
            csv_data.to_csv(csv_file_name, index=False)
